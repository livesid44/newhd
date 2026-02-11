#!/usr/bin/env python3
"""
Excel Wireframe Analyzer
Analyzes the 1234.xlsx wireframe file to extract visit tracker requirements
"""

import sys
import os

def check_file_exists(filepath):
    """Check if the Excel file exists"""
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        print("\nPlease add the 1234.xlsx file to:")
        print("  VisitManagement/wwwroot/1234.xlsx")
        return False
    print(f"✅ File found: {filepath}")
    return True

def analyze_excel(filepath):
    """Analyze the Excel wireframe file"""
    try:
        # Try to import openpyxl for Excel reading
        import openpyxl
        
        print(f"\n📊 Analyzing: {filepath}\n")
        
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        
        print(f"Workbook contains {len(workbook.sheetnames)} sheet(s):")
        for idx, sheet_name in enumerate(workbook.sheetnames, 1):
            print(f"  {idx}. {sheet_name}")
        
        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            print(f"\n{'='*60}")
            print(f"Sheet: {sheet_name}")
            print(f"{'='*60}")
            
            sheet = workbook[sheet_name]
            
            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"Dimensions: {max_row} rows × {max_col} columns")
            
            # Extract headers (first row)
            headers = []
            for col in range(1, min(max_col + 1, 50)):  # Limit to 50 columns
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            if headers:
                print(f"\nHeaders ({len(headers)} columns):")
                for idx, header in enumerate(headers, 1):
                    print(f"  {idx:2d}. {header}")
            
            # Show sample data (first 5 rows)
            if max_row > 1:
                print(f"\nSample data (first {min(max_row-1, 5)} rows):")
                for row_idx in range(2, min(max_row + 1, 7)):
                    row_data = []
                    for col in range(1, min(len(headers) + 1, max_col + 1)):
                        cell_value = sheet.cell(row=row_idx, column=col).value
                        row_data.append(str(cell_value) if cell_value else "")
                    if any(row_data):
                        print(f"  Row {row_idx}: {' | '.join(row_data[:5])}")  # Show first 5 columns
        
        # Focus on "Upcoming Visit tracker" sheet if it exists
        target_sheet = None
        for name in workbook.sheetnames:
            if 'upcoming' in name.lower() and 'visit' in name.lower():
                target_sheet = name
                break
        
        if target_sheet:
            print(f"\n{'='*60}")
            print(f"🎯 FOCUS: '{target_sheet}' Sheet Analysis")
            print(f"{'='*60}")
            sheet = workbook[target_sheet]
            
            # Extract all headers
            headers = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            print(f"\nAll Fields Required ({len(headers)}):")
            for idx, header in enumerate(headers, 1):
                print(f"  {idx:2d}. {header}")
            
            # Compare with current model
            print("\n📋 Field Mapping Notes:")
            print("  - Compare these fields with current Visit.cs model")
            print("  - Identify new fields to add")
            print("  - Identify fields to remove")
            print("  - Update form layouts accordingly")
        
        print(f"\n{'='*60}")
        print("Analysis complete!")
        print(f"{'='*60}\n")
        
    except ImportError:
        print("\n⚠️  openpyxl not installed. Installing...")
        os.system(f"{sys.executable} -m pip install openpyxl --quiet")
        print("✅ openpyxl installed. Please run the script again.")
        return
    except Exception as e:
        print(f"\n❌ Error analyzing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Default path to the Excel file
    default_path = "VisitManagement/wwwroot/1234.xlsx"
    
    filepath = sys.argv[1] if len(sys.argv) > 1 else default_path
    
    print("="*60)
    print("Excel Wireframe Analyzer for Visit Management System")
    print("="*60)
    
    if check_file_exists(filepath):
        analyze_excel(filepath)
    else:
        print("\n💡 Usage:")
        print(f"  python3 analyze_wireframe.py [path/to/excel/file]")
        print(f"\nDefault path: {default_path}")
        sys.exit(1)
